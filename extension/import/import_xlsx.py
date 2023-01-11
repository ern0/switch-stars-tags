import sys
import os
import openpyxl as xls


class Importer:

    def main(self):

        if False:
            #self.proc_url("https://www.nintendo.co.uk/Games/Nintendo-Switch-download-software/Trigonal-2161950.html")
            self.proc_url("https://www.nintendo.com/games/detail/the-bullet-time-of-revenge-switch/")
            quit()
        
        wb = xls.load_workbook(sys.argv[1])
        print("tags = {")        
        index = 0
        for name in wb.sheetnames:
            if name.endswith("Flips") or name.endswith("Arts"):
                sheet = wb[wb.sheetnames[index]]
                self.proc_sheet(sheet)
            index += 1
        print("}")
    
    def proc_sheet(self, sheet):
        
        print("    //", sheet.title)
        
        for eshop_column in range(1, 9):
            title = sheet.cell(column=eshop_column, row=1).value
            if title.startswith("Eshop"): break

        row = 2
        url = sheet.cell(column=eshop_column, row=row).value
        if url is None: row = 3
        
        while True:
            url = sheet.cell(column=eshop_column, row=row).value
            if url is None: break            
            self.proc_url(url)
            row += 1
        
    def proc_url(self, url):
        
        url = url.lower().strip()
        if url.endswith("/"): url = url[:-1]
        
        url = self.cut_end(url, "#overview")
        url = self.cut_end(url, "#gallery")
        url = self.cut_end(url, "-switch")
        url = self.cut_end(url, ".html")
        
        url = self.cut_number(url)

        name = os.path.basename(url)
        name = name.replace("-", "_")
        
s        
        print('    "' + name + '": "asset_flip",')
    
    def cut_end(self, text, pattern):
        
        if text.endswith(pattern):
            text = text[0:-len(pattern)]
        
        return text

    def cut_number(self, text):
        
        last = text.split("-")[-1]
        if len(last) > 5 and last.isnumeric():
            return text.rsplit("-", 1)[0]

        return text            


if __name__ == "__main__":
    (Importer()).main()
